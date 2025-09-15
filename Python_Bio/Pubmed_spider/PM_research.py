# 搜索excel里的PMID，并搜索GEO页面
import re
import os
from openpyxl import Workbook
from pathlib import Path
import numpy as np
import pandas as pd
import urllib.response
from bs4 import BeautifulSoup
import requests
from tqdm import tqdm
import requests
from tqdm import tqdm
import time
import sys


# -----------------------------------------------------------------------------------------------------------
def append_to_excel_pandas(data, filename, sheet_name_my):
    """
    使用 pandas 追加数据到 Excel 文件

    参数:
    - data: 要追加的数据，可以是字典、列表或 DataFrame
    - filename: Excel 文件名
    - sheet_name: 工作表名
    """
    # 如果文件不存在，创建新文件
    if not os.path.exists(filename):
        df = pd.DataFrame(data)
        df.to_excel(filename, sheet_name_my, index=False)
        print(f"创建新文件 {filename} 并写入数据")
    else:
        # 读取现有文件
        existing_df = pd.read_excel(filename, sheet_name_my)

        # 将新数据转换为 DataFrame
        if isinstance(data, dict):
            new_df = pd.DataFrame([data])
        elif isinstance(data, list):
            new_df = pd.DataFrame(data)
        else:
            new_df = data

        # 合并数据
        combined_df = pd.concat([existing_df, new_df], ignore_index=True)

        # 写回文件
        combined_df.to_excel(filename, sheet_name_my, index=False)
        # print(f"数据已追加到 {filename}")



def contains_any_fast(pattern,text):
    return bool(pattern.search(text))





# -----------------------------------------------------------------------------------------------------------









# -----------------------------------------------------------------------------------------------------------
print('Pubmed_research.py is running, let\'s go !!!')


input_file_path = './Pubmed_DATA/'
input_file_name = 'new_version.xlsx'
input_file = input_file_path + input_file_name
output_file_path = input_file_path
output_file_name = 'detail_'+input_file_name
# sheet_name = ['VEGF','DSCR1','FOXC2','CCBE1','GATA2','PIEZO1','VE_cadherin','TREM2','CD33','CD200','YKL_40','RIPK1','NLRP3','UBE2N']
sheet_name = ['DSCR1','FOXC2','PIEZO1']

# try:
#     # df_VEGF = pd.read_excel(input_file, sheet_name[0])
#     print(r'Complete: {} file read'.format(input_file_name))
# except IOError:
#     print(r'ERROR: {} file not read'.format(input_file_name))
#     sys.exit()


if not os.path.exists(output_file_path+output_file_name):
    # 创建工作簿
    wb = Workbook()
    # 删除默认创建的工作表（如果需要）
    if 'Sheet' in wb.sheetnames:
        std_sheet = wb['Sheet']
        wb.remove(std_sheet)
    for num in range(len(sheet_name)):
        sheet_name_each = sheet_name[num]
        wb.create_sheet(sheet_name_each)
    wb.save(output_file_path+output_file_name)
    print(f"创建新文件 {output_file_name}")



for iter_num, item_iter in tqdm(enumerate(sheet_name), total=len(sheet_name), desc="Factor Processing"):
    df_temp = pd.read_excel(input_file, sheet_name[iter_num])
    column_data = df_temp['PMID']

    data_list = []
    for i, item in tqdm(enumerate(column_data), total=len(column_data), desc="PMID Processing", leave=False):
        url = 'https://pubmed.ncbi.nlm.nih.gov/' + str(column_data[i]) + '/'

        response = requests.get(url)
        html_content = response.text
        # print(html_content)
        try:
            pattern_Full_text_links = re.compile('<h3 class="title">.*?Full text links.*?</h3>(.*?)</div>', re.S)
            data_Full_text_links = re.findall(pattern_Full_text_links, html_content)
            pattern_links = re.compile('   href="(.*?)".*?target="_blank"', re.S)
            data_links = re.findall(pattern_links, str(data_Full_text_links))
            data_links = list(set(data_links))
            exist_Methods = ''
            for num, item in enumerate(data_links):
                response_temp = requests.get(data_links[num])
                html_temp = response_temp.text

                keywords = ["Human", "Patient"]
                keywords_pattern = re.compile('|'.join(map(re.escape, keywords)))
                keywords_Methods = contains_any_fast(keywords_pattern,html_temp)
                if keywords_Methods:
                    exist_Methods = exist_Methods + 'Human,'

                keywords = ["Enzyme immunoassay", "enzyme immunoassay", "ELISA","immunoassay","enzyme-linked", "immunosorbent assay"]
                keywords_pattern = re.compile('|'.join(map(re.escape, keywords)))
                keywords_Methods = contains_any_fast(keywords_pattern,html_temp)
                if keywords_Methods:
                    exist_Methods = exist_Methods + 'ELISA,'

                keywords = ["PCR", "Polymerase Chain Reaction","Polymerase chain reaction","polymerase chain reaction","mRNA","siRNA","shRNA"]
                keywords_pattern = re.compile('|'.join(map(re.escape, keywords)))
                keywords_Methods = contains_any_fast(keywords_pattern, html_temp)
                if keywords_Methods:
                    exist_Methods = exist_Methods + 'PCR,'

                keywords = ["WB", "Western blotting", "Western Blotting", "western blotting","Immunoblotting","immunoblotting"]
                keywords_pattern = re.compile('|'.join(map(re.escape, keywords)))
                keywords_Methods = contains_any_fast(keywords_pattern, html_temp)
                if keywords_Methods:
                    exist_Methods = exist_Methods + 'WB,'

                keywords = ["flow cytometry", "Flow cytometry"]
                keywords_pattern = re.compile('|'.join(map(re.escape, keywords)))
                keywords_Methods = contains_any_fast(keywords_pattern, html_temp)
                if keywords_Methods:
                    exist_Methods = exist_Methods + 'flow cytometry,'

                keywords = ["Exome sequence", "whole-exome", "exome sequencing","genomic","genome","Whole-exome","genome-wide"]
                keywords_pattern = re.compile('|'.join(map(re.escape, keywords)))
                keywords_Methods = contains_any_fast(keywords_pattern, html_temp)
                if keywords_Methods:
                    exist_Methods = exist_Methods + 'WES,'

                # BioLegend, eBiosciences, BD Biosciences, Fermentas, Qiagen, Sigma,


            if not exist_Methods:
                exist_Methods = 'not exist'
            temp_data = [[column_data[i],exist_Methods]]
        except IndexError:
            temp_data = [[column_data[i],'not find PMID']]
        except requests.exceptions.ConnectTimeout:
            temp_data = [[column_data[i],'url ConnectTimeout']]
        except requests.exceptions.ReadTimeout:
            temp_data = [[column_data[i],'url ReadTimeout']]
        except requests.exceptions.RequestException as e:
            temp_data = [[column_data[i],'url RequestException']]

        data_list.extend(temp_data)
    print(data_list)

    append_to_excel_pandas(data_list,output_file_path+sheet_name[iter_num]+'_'+output_file_name, sheet_name[iter_num])




